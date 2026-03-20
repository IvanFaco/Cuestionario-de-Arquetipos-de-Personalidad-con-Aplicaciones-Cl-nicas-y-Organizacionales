import streamlit as st
import pandas as pd
import plotly.express as px
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF
import base64

# --- CONFIGURACIÓN DE LA PÁGINA ---
st.set_page_config(page_title="Cuestionario de Arquetipos", layout="wide", initial_sidebar_state="collapsed")

# --- CUSTOM STYLING ---
st.markdown("""
<style>
    body {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    }
    .main {
        background-color: #f8f9fa;
        border-radius: 10px;
    }
    h1 {
        color: #667eea;
        text-align: center;
        font-size: 2.5em;
        margin-bottom: 10px;
    }
    h2 {
        color: #764ba2;
        border-bottom: 3px solid #667eea;
        padding-bottom: 10px;
    }
    .stExpander {
        background-color: #ffffff;
        border-radius: 8px;
        border-left: 4px solid #667eea;
    }
    .stRadio {
        display: flex;
        gap: 10px;
        padding: 10px 0;
    }
    .question-text {
        font-size: 1.05em;
        font-weight: 500;
        color: #2c3e50;
        margin-bottom: 10px;
    }
</style>
""", unsafe_allow_html=True)

st.title("Cuestionario de Evaluación Integral")
st.markdown("### Instrucciones")
st.markdown("""
Responde las siguientes afirmaciones según tu experiencia personal. 
Selecciona tu nivel de acuerdo en una escala de **Totalmente en Desacuerdo** a **Totalmente de Acuerdo**.
Todas tus respuestas son confidenciales y se procesan de forma anónima.
""")
st.divider()

# --- BANCO DE PREGUNTAS COMPLETO (60 ÍTEMS) ---
preguntas_roles = {
    "Inocente": [
        "Confío en que las cosas saldrán bien al final, a pesar de los problemas.", 
        "Intento ver lo bueno en todas las personas y situaciones.",
        "Me siento seguro/a y protegido/a en mi entorno actual.",
        "Suelo mantener una actitud optimista ante el futuro."
    ],
    "Huérfano": [
        "Soy realista y sé que la vida es dura y a menudo injusta.", 
        "Siento que debo valerme por mí mismo/a sin depender de otros.",
        "Me preparo constantemente para que las cosas salgan mal.",
        "He aprendido que nadie vendrá a rescatarme de mis problemas."
    ],
    "Guerrero": [
        "Me motivan los desafíos y competir para ganar.", 
        "Defiendo mis posturas con firmeza y no me rindo fácilmente.",
        "Soy capaz de tomar decisiones difíciles y ejecutar planes de acción.",
        "Me siento vivo/a cuando supero obstáculos difíciles."
    ],
    "Cuidador": [
        "Ayudar y proteger a los demás es mi principal prioridad.", 
        "Me cuesta decir 'no' cuando alguien necesita apoyo o consuelo.",
        "Siento una gran satisfacción al nutrir el crecimiento de otras personas.",
        "A menudo pongo las necesidades de los demás por encima de las mías."
    ],
    "Explorador": [
        "Necesito libertad para descubrir cosas nuevas y viajar.", 
        "Me aburre la rutina y busco aventuras o experiencias inéditas.",
        "Siento un fuerte impulso por descubrir quién soy realmente.",
        "Prefiero la independencia antes que sentirme atado/a a compromisos fijos."
    ],
    "Amante": [
        "Busco conexiones emocionales profundas e intimidad genuina.", 
        "La estética, el arte y la pasión son fundamentales para mí.",
        "Me esfuerzo por crear armonía y belleza en mis relaciones y entorno.",
        "Tiendo a entregarme por completo a lo que amo (personas, proyectos, ideas)."
    ],
    "Rebelde": [
        "Cuestiono las reglas, tradiciones o autoridades que me parecen injustas.", 
        "No me importa romper las normas si considero que es lo correcto.",
        "Siento la necesidad de destruir lo que ya no funciona para empezar de nuevo.",
        "A menudo me siento como un/a forastero/a que desafía el status quo."
    ],
    "Creador": [
        "Siento una necesidad constante de innovar, diseñar o construir algo nuevo.", 
        "Me expreso mejor a través de mis creaciones e imaginación.",
        "Me frustra cuando no puedo darle forma a las ideas que tengo en la cabeza.",
        "Busco dejar un legado tangible o una marca única en el mundo."
    ],
    "Gobernante": [
        "Me gusta tomar el control y organizar a los equipos o recursos.", 
        "Asumo la responsabilidad del éxito o fracaso de los proyectos.",
        "Me siento cómodo/a estableciendo reglas, estructuras y orden.",
        "Creo que el liderazgo fuerte es clave para evitar el caos."
    ],
    "Mago": [
        "Creo en transformar la realidad cambiando primero la mentalidad.", 
        "Me interesan los procesos de sanación y transformación profunda.",
        "A menudo tengo intuiciones o visiones sobre cómo resultarán las cosas.",
        "Me fascina entender cómo funcionan las leyes ocultas de la mente o el universo."
    ],
    "Sabio": [
        "Busco la verdad objetiva basándome en datos y análisis racional.", 
        "Prefiero analizar fríamente una situación antes de actuar.",
        "Valoro el conocimiento y el aprendizaje continuo por encima de todo.",
        "Intento ser un observador neutral y no dejarme llevar por las emociones."
    ],
    "Bufón": [
        "Uso el humor para aligerar situaciones tensas o incómodas.", 
        "Creo que lo más importante es disfrutar el momento presente.",
        "No me tomo la vida ni a mí mismo/a demasiado en serio.",
        "Me gusta ser el centro de atención y hacer reír a los demás."
    ]
}

preguntas_estructura = {
    "Persona": [
        "Me esfuerzo mucho por cuidar la imagen que proyecto en lo profesional.", 
        "A veces siento que uso una 'máscara' social para encajar.",
        "Me preocupa profundamente lo que la sociedad o mis colegas piensen de mí."
    ],
    "Sombra": [
        "Tengo pensamientos, impulsos o deseos que me avergüenza admitir.", 
        "A menudo me critico duramente en secreto.",
        "Hay partes de mi personalidad que oculto porque creo que serían rechazadas."
    ],
    "Anima_Animus": [
        "Equilibrio bien mi lado emocional con mi lado lógico y asertivo.", 
        "Me conecto fácilmente con mi intuición y mis sentimientos más profundos.",
        "Soy capaz de ser tanto receptivo/a y empático/a como firme y directo/a según se requiera."
    ],
    "Self": [
        "Siento que mi vida tiene un propósito claro y un sentido profundo.", 
        "Experimento una gran coherencia interna entre lo que pienso, digo y hago.",
        "Me siento conectado/a con algo más grande que mi propio ego."
    ]
}

# --- FUNCIÓN PARA GENERAR ANÁLISIS DETALLADO ---
def generar_analisis(promedios_roles, promedios_est):
    """Genera el análisis detallado basado en los promedios"""
    persona_score = promedios_est["Persona"]
    sombra_score = promedios_est["Sombra"]
    self_score = promedios_est["Self"]
    
    ordenados = sorted(promedios_roles.items(), key=lambda x: x[1], reverse=True)
    dominantes = ordenados[:3]
    
    triada = f"Tu tríada de liderazgo psíquico:\n"
    triada += f"1. {dominantes[0][0]} (Puntuación: {dominantes[0][1]:.2f}/5.0)\n"
    triada += f"2. {dominantes[1][0]} (Puntuación: {dominantes[1][1]:.2f}/5.0)\n"
    triada += f"3. {dominantes[2][0]} (Puntuación: {dominantes[2][1]:.2f}/5.0)"
    
    if persona_score > 4.0 and sombra_score < 2.5:
        diagnostico = f"ALERTA: Posible Agotamiento Social\n\nTu Persona está muy elevada ({persona_score:.2f}) mientras tu Sombra es baja ({sombra_score:.2f}). Esto sugiere que inviertes mucha energía en mantener una imagen pública impecable, posiblemente reprimiendo aspectos naturales de tu personalidad. Considera espacios seguros para ser más auténtico/a."
    elif sombra_score >= 3.5:
        diagnostico = f"ALERTA: Presión del Inconsciente\n\nTu puntuación en Sombra es elevada ({sombra_score:.2f}). Hay partes de ti que demandan atención: impulsos reprimidos, deseos ocultos o conflictos internos. Una reflexión profunda o acompañamiento profesional podría ayudarte a integrar estos aspectos."
    elif self_score < 2.5:
        diagnostico = f"ALERTA: Crisis de Sentido\n\nTu puntuación en Self es baja ({self_score:.2f}). Podrías estar experimentando una desconexión con tu propósito vital o identidad profunda. Actividades reflexivas o exploratorias podrían ayudarte a reconectar con tu esencia."
    else:
        diagnostico = f"ESTADO: Balance Psicológico Funcional\n\nTu relación entre Persona ({persona_score:.2f}), Sombra ({sombra_score:.2f}) y Self ({self_score:.2f}) demuestra un buen equilibrio. Estás integrando tus aspectos públicos, internos y tu sentido de propósito de manera relativamente armoniosa."
    
    return triada, diagnostico

# --- FUNCIÓN PARA GENERAR EXCEL CON RESULTADOS ---
def generar_excel_reporte(respuestas, promedios_roles, promedios_est):
    """Crea un archivo Excel con datos, gráficos y análisis"""
    
    # Crear un workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remover la hoja por defecto
    
    # --- HOJA 1: DATOS CRUDOS ---
    ws_datos = wb.create_sheet("Datos Crudos", 0)
    ws_datos['A1'] = "Cuestionario de Arquetipos de Personalidad"
    ws_datos['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_datos['A1'].fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    ws_datos.merge_cells('A1:F1')
    
    ws_datos['A2'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    
    # Datos de arquetipos
    row = 4
    ws_datos['A4'] = "ARQUETIPOS DE PERSONALIDAD"
    ws_datos['A4'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_datos['A4'].fill = PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid")
    
    row = 5
    for arquetipo, valores in respuestas.items():
        if arquetipo in [k for k in promedios_roles.keys()]:
            ws_datos[f'A{row}'] = arquetipo
            for col_idx, valor in enumerate(valores, 1):
                ws_datos.cell(row=row, column=col_idx+1, value=valor)
            row += 1
    
    # Datos de dimensiones
    row += 2
    ws_datos[f'A{row}'] = "DIMENSIONES ESTRUCTURALES"
    ws_datos[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_datos[f'A{row}'].fill = PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid")
    
    row += 1
    for dim, valores in respuestas.items():
        if dim in [k for k in promedios_est.keys()]:
            ws_datos[f'A{row}'] = dim.replace('_', ' / ')
            for col_idx, valor in enumerate(valores, 1):
                ws_datos.cell(row=row, column=col_idx+1, value=valor)
            row += 1
    
    # Ajustar ancho de columnas
    ws_datos.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws_datos.column_dimensions[col].width = 12
    
    # --- HOJA 2: PROMEDIOS Y ANÁLISIS ---
    ws_analisis = wb.create_sheet("Análisis Detallado", 1)
    ws_analisis['A1'] = "RESULTADOS Y ANÁLISIS"
    ws_analisis['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_analisis['A1'].fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    ws_analisis.merge_cells('A1:C1')
    
    # Tabla de promedios de arquetipos
    row = 3
    ws_analisis[f'A{row}'] = "ARQUETIPOS - PUNTUACIONES"
    ws_analisis[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws_analisis[f'A{row}'].fill = PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid")
    ws_analisis[f'B{row}'].fill = PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid")
    
    row += 1
    ws_analisis[f'A{row}'] = "Dimensión"
    ws_analisis[f'B{row}'] = "Puntuación"
    ws_analisis[f'A{row}'].font = Font(bold=True)
    ws_analisis[f'B{row}'].font = Font(bold=True)
    
    row += 1
    ordenados = sorted(promedios_roles.items(), key=lambda x: x[1], reverse=True)
    for arch, puntaje in ordenados:
        ws_analisis[f'A{row}'] = arch
        ws_analisis[f'B{row}'] = round(puntaje, 2)
        row += 1
    
    # Tabla de dimensiones estructurales
    row += 2
    ws_analisis[f'A{row}'] = "ESTRUCTURA PSICOLÓGICA"
    ws_analisis[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws_analisis[f'A{row}'].fill = PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid")
    ws_analisis[f'B{row}'].fill = PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid")
    
    row += 1
    ws_analisis[f'A{row}'] = "Dimensión"
    ws_analisis[f'B{row}'] = "Puntuación"
    ws_analisis[f'A{row}'].font = Font(bold=True)
    ws_analisis[f'B{row}'].font = Font(bold=True)
    
    row += 1
    for dim, puntaje in promedios_est.items():
        ws_analisis[f'A{row}'] = dim.replace('_', ' / ')
        ws_analisis[f'B{row}'] = round(puntaje, 2)
        row += 1
    
    # Diagnóstico
    row += 2
    ws_analisis[f'A{row}'] = "DIAGNÓSTICO Y RECOMENDACIONES"
    ws_analisis[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws_analisis[f'A{row}'].fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    ws_analisis.merge_cells(f'A{row}:B{row}')
    
    row += 1
    persona_score = promedios_est["Persona"]
    sombra_score = promedios_est["Sombra"]
    self_score = promedios_est["Self"]
    
    dominantes = ordenados[:3]
    
    triada, diagnostico = generar_analisis(promedios_roles, promedios_est)
    
    ws_analisis[f'A{row}'] = triada
    ws_analisis[f'A{row}'].alignment = Alignment(wrap_text=True)
    ws_analisis.merge_cells(f'A{row}:B{row}')
    ws_analisis.row_dimensions[row].height = 60
    
    row += 4
    ws_analisis[f'A{row}'] = diagnostico
    ws_analisis[f'A{row}'].alignment = Alignment(wrap_text=True)
    ws_analisis[f'A{row}'].font = Font(size=10)
    if "ALERTA" in diagnostico:
        ws_analisis[f'A{row}'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    else:
        ws_analisis[f'A{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws_analisis.merge_cells(f'A{row}:B{row}')
    ws_analisis.row_dimensions[row].height = 80
    
    ws_analisis.column_dimensions['A'].width = 40
    ws_analisis.column_dimensions['B'].width = 15
    
    # Guardar en BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

# --- FUNCIÓN PARA GENERAR PDF CON RESULTADOS ---
def generar_pdf_reporte(respuestas, promedios_roles, promedios_est, fig_radar, fig_bar):
    """Crea un archivo PDF con análisis, gráficos y resultados"""
    from fpdf import FPDF
    
    # Convertir gráficos plotly a imágenes
    try:
        # Crear imágenes temporales de los gráficos
        img_radar_bytes = fig_radar.to_image(format="png", width=800, height=600)
        img_bar_bytes = fig_bar.to_image(format="png", width=800, height=600)
        
        # Guardar en BytesIO
        radar_tmp = BytesIO(img_radar_bytes)
        bar_tmp = BytesIO(img_bar_bytes)
    except:
        radar_tmp = None
        bar_tmp = None
    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=12)
    
    # Encabezado
    pdf.set_font("Helvetica", "B", size=16)
    pdf.cell(0, 10, "Cuestionario de Arquetipos de Personalidad", ln=True, align="C")
    pdf.set_font("Helvetica", size=10)
    pdf.cell(0, 10, f"Evaluación generada: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True, align="C")
    pdf.ln(5)
    
    # Obtener análisis
    triada, diagnostico = generar_analisis(promedios_roles, promedios_est)
    
    # Sección de Tríada
    pdf.set_font("Helvetica", "B", size=12)
    pdf.cell(0, 10, "Tríada de Liderazgo Psíquico", ln=True)
    pdf.set_font("Helvetica", size=10)
    pdf.multi_cell(0, 5, triada)
    pdf.ln(3)
    
    # Tabla de Arquetipos
    pdf.set_font("Helvetica", "B", size=12)
    pdf.cell(0, 10, "Puntuaciones de Arquetipos", ln=True)
    pdf.set_font("Helvetica", size=9)
    
    ordenados = sorted(promedios_roles.items(), key=lambda x: x[1], reverse=True)
    col_width = pdf.w / 3
    
    pdf.set_fill_color(102, 126, 234)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(col_width * 1.5, 7, "Arquetipo", border=1, fill=True, align="C")
    pdf.cell(col_width * 1.5, 7, "Puntuación", border=1, fill=True, align="C", ln=True)
    
    pdf.set_text_color(0, 0, 0)
    for arch, puntaje in ordenados:
        pdf.cell(col_width * 1.5, 7, arch, border=1, align="L")
        pdf.cell(col_width * 1.5, 7, f"{puntaje:.2f}/5.0", border=1, align="C", ln=True)
    
    pdf.ln(5)
    
    # Tabla de Dimensiones
    pdf.set_font("Helvetica", "B", size=12)
    pdf.cell(0, 10, "Estructura Psicológica", ln=True)
    pdf.set_font("Helvetica", size=9)
    
    pdf.set_fill_color(102, 126, 234)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(col_width * 1.5, 7, "Dimensión", border=1, fill=True, align="C")
    pdf.cell(col_width * 1.5, 7, "Puntuación", border=1, fill=True, align="C", ln=True)
    
    pdf.set_text_color(0, 0, 0)
    for dim, puntaje in promedios_est.items():
        pdf.cell(col_width * 1.5, 7, dim.replace("_", " / "), border=1, align="L")
        pdf.cell(col_width * 1.5, 7, f"{puntaje:.2f}/5.0", border=1, align="C", ln=True)
    
    # Agregar gráficos si están disponibles
    if radar_tmp and bar_tmp:
        pdf.add_page()
        pdf.set_font("Helvetica", "B", size=14)
        pdf.cell(0, 10, "Gráficos de Resultados", ln=True)
        pdf.ln(5)
        
        # Gráfico radar
        pdf.image(radar_tmp, x=10, y=pdf.get_y(), w=180)
        pdf.ln(100)
        
        # Nuevo página para el bar
        pdf.add_page()
        pdf.set_font("Helvetica", "B", size=14)
        pdf.cell(0, 10, "Estructura Psicológica", ln=True)
        pdf.ln(5)
        pdf.image(bar_tmp, x=10, y=pdf.get_y(), w=180)
        pdf.ln(100)
    
    # Nueva página para análisis
    pdf.add_page()
    pdf.set_font("Helvetica", "B", size=14)
    pdf.cell(0, 10, "Diagnóstico y Análisis Detallado", ln=True)
    pdf.ln(5)
    
    pdf.set_font("Helvetica", size=10)
    pdf.multi_cell(0, 5, diagnostico)
    
    pdf.ln(5)
    pdf.set_font("Helvetica", "I", size=8)
    pdf.multi_cell(0, 4, "Nota: Este instrumento sigue principios de análisis Junguianos. Los resultados son tendencias narrativas, no diagnósticos psiquiátricos concluyentes.")
    
    # Guardar en BytesIO
    pdf_buffer = BytesIO()
    pdf_bytes = pdf.output()
    pdf_buffer.write(pdf_bytes)
    pdf_buffer.seek(0)
    
    return pdf_buffer

# --- INTERFAZ DEL CUESTIONARIO ---
respuestas = {}
contador_seccion = 1

# Crear lista de mapeo para arquetipos con números de sección
seccion_a_arquetipo = {}
for idx, (arquetipo, _) in enumerate(preguntas_roles.items(), 1):
    seccion_a_arquetipo[idx] = arquetipo

# Inicializar session_state para respuestas en tiempo real
if "respuestas_tiempo_real" not in st.session_state:
    st.session_state.respuestas_tiempo_real = {}

if "seccion_activa" not in st.session_state:
    st.session_state.seccion_activa = None

def actualizar_respuesta(clave, valor, seccion_num):
    """Callback para actualizar respuestas en tiempo real y marcar sección activa"""
    st.session_state.respuestas_tiempo_real[clave] = valor
    st.session_state.seccion_activa = seccion_num

def obtener_porcentaje_seccion(seccion_num, preguntas_dict, es_estructura=False):
    """Calcula el porcentaje de gestión de una sección"""
    prefix = "E" if es_estructura else "R"
    
    if es_estructura:
        dims = list(preguntas_dict.keys())
        dim = dims[seccion_num - 13] if seccion_num >= 13 else None
        if not dim:
            return 0
        num_preguntas = len(preguntas_dict[dim])
    else:
        arquetipos = list(preguntas_dict.keys())
        arquetipo = arquetipos[seccion_num - 1] if seccion_num <= len(arquetipos) else None
        if not arquetipo:
            return 0
        num_preguntas = len(preguntas_dict[arquetipo])
    
    respuestas_en_seccion = sum(1 for k in st.session_state.respuestas_tiempo_real.keys() 
                               if k.startswith(f"{prefix}_{seccion_num}_"))
    
    return int((respuestas_en_seccion / num_preguntas) * 100) if num_preguntas > 0 else 0

def obtener_color_indicador(porcentaje):
    """Retorna el emoji indicador según el porcentaje"""
    if porcentaje == 0:
        return "🔴"
    elif porcentaje < 50:
        return "🟠"
    elif porcentaje < 100:
        return "🟡"
    else:
        return "🟢"

st.header("Parte I: Preguntas generales")
col1, col2 = st.columns(2)

with col1:
    for arquetipo, lista_preg in list(preguntas_roles.items())[::2]:
        seccion_num = list(preguntas_roles.keys()).index(arquetipo) + 1
        porcentaje = obtener_porcentaje_seccion(seccion_num, preguntas_roles)
        color_indicador = obtener_color_indicador(porcentaje)
        
        # Solo expandir si es la sección activa
        esta_abierta = st.session_state.seccion_activa == seccion_num
        
        with st.expander(f"Sección {seccion_num} {color_indicador} ({porcentaje}%)", expanded=esta_abierta):
            respuestas[arquetipo] = []
            for i, preg in enumerate(lista_preg):
                st.markdown(f"<p class='question-text'>{preg}</p>", unsafe_allow_html=True)
                clave = f"R_{seccion_num}_{i}"
                val = st.radio(
                    label="Tu respuesta:",
                    options=[1, 2, 3, 4, 5],
                    format_func=lambda x: {
                        1: "Totalmente en Desacuerdo",
                        2: "En Desacuerdo",
                        3: "Neutral",
                        4: "De Acuerdo",
                        5: "Totalmente de Acuerdo"
                    }[x],
                    key=clave,
                    horizontal=True,
                    on_change=actualizar_respuesta,
                    args=(clave, st.session_state.get(clave, 3), seccion_num)
                )
                respuestas[arquetipo].append(val)
                    
with col2:
    for arquetipo, lista_preg in list(preguntas_roles.items())[1::2]:
        seccion_num = list(preguntas_roles.keys()).index(arquetipo) + 1
        porcentaje = obtener_porcentaje_seccion(seccion_num, preguntas_roles)
        color_indicador = obtener_color_indicador(porcentaje)
        
        # Solo expandir si es la sección activa
        esta_abierta = st.session_state.seccion_activa == seccion_num
        
        with st.expander(f"Sección {seccion_num} {color_indicador} ({porcentaje}%)", expanded=esta_abierta):
            respuestas[arquetipo] = []
            for i, preg in enumerate(lista_preg):
                st.markdown(f"<p class='question-text'>{preg}</p>", unsafe_allow_html=True)
                clave = f"R_{seccion_num}_{i}"
                val = st.radio(
                    label="Tu respuesta:",
                    options=[1, 2, 3, 4, 5],
                    format_func=lambda x: {
                        1: "Totalmente en Desacuerdo",
                        2: "En Desacuerdo",
                        3: "Neutral",
                        4: "De Acuerdo",
                        5: "Totalmente de Acuerdo"
                    }[x],
                    key=clave,
                    horizontal=True,
                    on_change=actualizar_respuesta,
                    args=(clave, st.session_state.get(clave, 3), seccion_num)
                )
                respuestas[arquetipo].append(val)

st.divider()
st.header("Parte II: Preguntas complementarias")

col3, col4 = st.columns(2)

with col3:
    for idx, (dim, lista_preg) in enumerate(list(preguntas_estructura.items())[::2]):
        seccion_num = 13 + (idx * 2)
        porcentaje = obtener_porcentaje_seccion(seccion_num, preguntas_estructura, es_estructura=True)
        color_indicador = obtener_color_indicador(porcentaje)
        
        # Solo expandir si es la sección activa
        esta_abierta = st.session_state.seccion_activa == seccion_num
        
        with st.expander(f"Sección {seccion_num} {color_indicador} ({porcentaje}%)", expanded=esta_abierta):
            respuestas[dim] = []
            for i, preg in enumerate(lista_preg):
                st.markdown(f"<p class='question-text'>{preg}</p>", unsafe_allow_html=True)
                clave = f"E_{seccion_num}_{i}"
                val = st.radio(
                    label="Tu respuesta:",
                    options=[1, 2, 3, 4, 5],
                    format_func=lambda x: {
                        1: "Totalmente en Desacuerdo",
                        2: "En Desacuerdo",
                        3: "Neutral",
                        4: "De Acuerdo",
                        5: "Totalmente de Acuerdo"
                    }[x],
                    key=clave,
                    horizontal=True,
                    on_change=actualizar_respuesta,
                    args=(clave, st.session_state.get(clave, 3), seccion_num)
                )
                respuestas[dim].append(val)
                    
with col4:
    for idx, (dim, lista_preg) in enumerate(list(preguntas_estructura.items())[1::2]):
        seccion_num = 14 + (idx * 2)
        porcentaje = obtener_porcentaje_seccion(seccion_num, preguntas_estructura, es_estructura=True)
        color_indicador = obtener_color_indicador(porcentaje)
        
        # Solo expandir si es la sección activa
        esta_abierta = st.session_state.seccion_activa == seccion_num
        
        with st.expander(f"Sección {seccion_num} {color_indicador} ({porcentaje}%)", expanded=esta_abierta):
            respuestas[dim] = []
            for i, preg in enumerate(lista_preg):
                st.markdown(f"<p class='question-text'>{preg}</p>", unsafe_allow_html=True)
                clave = f"E_{seccion_num}_{i}"
                val = st.radio(
                    label="Tu respuesta:",
                    options=[1, 2, 3, 4, 5],
                    format_func=lambda x: {
                        1: "Totalmente en Desacuerdo",
                        2: "En Desacuerdo",
                        3: "Neutral",
                        4: "De Acuerdo",
                        5: "Totalmente de Acuerdo"
                    }[x],
                    key=clave,
                    horizontal=True,
                    on_change=actualizar_respuesta,
                    args=(clave, st.session_state.get(clave, 3), seccion_num)
                )
                respuestas[dim].append(val)

st.divider()

col_button = st.columns([1, 4])[0]
with col_button:
    submit = st.button("📊 Generar Análisis Completo", use_container_width=True)

# Usar session_state para mantener los resultados
if "mostrar_resultados" not in st.session_state:
    st.session_state.mostrar_resultados = False

if "respuestas_guardadas" not in st.session_state:
    st.session_state.respuestas_guardadas = None

if "promedios_roles_guardados" not in st.session_state:
    st.session_state.promedios_roles_guardados = None

if "promedios_est_guardados" not in st.session_state:
    st.session_state.promedios_est_guardados = None

if "fig_radar_guardado" not in st.session_state:
    st.session_state.fig_radar_guardado = None

if "fig_bar_guardado" not in st.session_state:
    st.session_state.fig_bar_guardado = None

# --- MOTOR DE ANÁLISIS ---
if submit:
    # Recopilar todas las respuestas del session_state
    respuestas_finales = {}
    
    # Recopilar arquetipos
    for idx, (arquetipo, lista_preg) in enumerate(preguntas_roles.items(), 1):
        respuestas_finales[arquetipo] = []
        for i in range(len(lista_preg)):
            clave = f"R_{idx}_{i}"
            valor = st.session_state.get(clave, 3)
            respuestas_finales[arquetipo].append(valor)
    
    # Recopilar dimensiones estructurales
    seccion_counter = 13
    for idx, (dim, lista_preg) in enumerate(preguntas_estructura.items()):
        if idx % 2 == 0:
            seccion_num = seccion_counter
            seccion_counter += 2
        else:
            seccion_num = seccion_counter - 1
        
        respuestas_finales[dim] = []
        for i in range(len(lista_preg)):
            clave = f"E_{seccion_num}_{i}"
            valor = st.session_state.get(clave, 3)
            respuestas_finales[dim].append(valor)
    
    # Guardar en session_state
    st.session_state.respuestas_guardadas = respuestas_finales
    st.session_state.promedios_roles_guardados = {k: sum(v)/len(v) for k, v in respuestas_finales.items() if k in preguntas_roles}
    st.session_state.promedios_est_guardados = {k: sum(v)/len(v) for k, v in respuestas_finales.items() if k in preguntas_estructura}
    
    # Crear gráficos
    promedios_roles = st.session_state.promedios_roles_guardados
    promedios_est = st.session_state.promedios_est_guardados
    
    df_radar = pd.DataFrame(dict(Puntaje=list(promedios_roles.values()), Dimensión=list(promedios_roles.keys())))
    fig_radar = px.line_polar(df_radar, r='Puntaje', theta='Dimensión', line_close=True, range_r=[0,5], markers=True)
    fig_radar.update_traces(fill='toself', line_color='#667eea', fillcolor='rgba(102, 126, 234, 0.4)')
    fig_radar.update_layout(
        polar=dict(bgcolor="rgba(240, 240, 250, 0.5)"),
        font=dict(size=10),
        height=500
    )
    
    df_bar = pd.DataFrame(dict(Dimensión=list(promedios_est.keys()), Puntaje=list(promedios_est.values())))
    fig_bar = px.bar(df_bar, x='Dimensión', y='Puntaje', range_y=[0,5], 
                     color='Puntaje', color_continuous_scale='Viridis')
    fig_bar.add_hline(y=3.0, line_dash="dot", line_color="red", annotation_text="Punto Neutro (3.0)")
    fig_bar.update_layout(height=500, showlegend=False)
    
    st.session_state.fig_radar_guardado = fig_radar
    st.session_state.fig_bar_guardado = fig_bar
    st.session_state.mostrar_resultados = True

# Mostrar resultados si están guardados
if st.session_state.mostrar_resultados:
    promedios_roles = st.session_state.promedios_roles_guardados
    promedios_est = st.session_state.promedios_est_guardados
    fig_radar = st.session_state.fig_radar_guardado
    fig_bar = st.session_state.fig_bar_guardado
    respuestas_finales = st.session_state.respuestas_guardadas
    
    st.divider()
    st.markdown("## 📈 Resultados de tu Evaluación")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("🎭 Perfil Personalidad (12 Dimensiones)")
        st.plotly_chart(fig_radar, use_container_width=True)
        
    with col2:
        st.subheader("🧠 Estructura Psicológica Integrada")
        st.plotly_chart(fig_bar, use_container_width=True)

    # --- GENERACIÓN DE DIAGNÓSTICO TEXTUAL ---
    st.markdown("---")
    st.subheader("📋 Análisis Detallado")
    
    ordenados = sorted(promedios_roles.items(), key=lambda x: x[1], reverse=True)
    dominantes = ordenados[:3]
    
    col_d1, col_d2, col_d3 = st.columns(3)
    with col_d1:
        st.metric("🥇 Dominante Principal", dominantes[0][0], f"{dominantes[0][1]:.2f}/5")
    with col_d2:
        st.metric("🥈 Secundario", dominantes[1][0], f"{dominantes[1][1]:.2f}/5")
    with col_d3:
        st.metric("🥉 Terciario", dominantes[2][0], f"{dominantes[2][1]:.2f}/5")
    
    st.markdown(f"""
    **Tu tríada de liderazgo psíquico:**
    
    1. **{dominantes[0][0]}** (Puntuación: {dominantes[0][1]:.2f}/5.0)
    2. **{dominantes[1][0]}** (Puntuación: {dominantes[1][1]:.2f}/5.0)
    3. **{dominantes[2][0]}** (Puntuación: {dominantes[2][1]:.2f}/5.0)
    """)
    
    persona_score = promedios_est["Persona"]
    sombra_score = promedios_est["Sombra"]
    self_score = promedios_est["Self"]
    
    st.markdown("### ⚖️ Síntesis de Integración Estructural")
    
    if persona_score > 4.0 and sombra_score < 2.5:
        st.warning(f"""
        **⚠️ Posible Agotamiento Social**
        
        Tu Persona está muy elevada ({persona_score:.2f}) mientras tu Sombra es baja ({sombra_score:.2f}). 
        Esto sugiere que inviertes mucha energía en mantener una imagen pública impecable, posiblemente reprimiendo 
        aspectos naturales de tu personalidad. Considera espacios seguros para ser más auténtico/a.
        """)
    elif sombra_score >= 3.5:
        st.error(f"""
        **🔴 Presión del Inconsciente**
        
        Tu puntuación en Sombra es elevada ({sombra_score:.2f}). Hay partes de ti que demandan atención: 
        impulsos reprimidos, deseos ocultos o conflictos internos. Una reflexión profunda o acompañamiento 
        profesional podría ayudarte a integrar estos aspectos.
        """)
    elif self_score < 2.5:
        st.warning(f"""
        **❓ Crisis de Sentido**
        
        Tu puntuación en Self es baja ({self_score:.2f}). Podrías estar experimentando una desconexión 
        con tu propósito vital o identidad profunda. Actividades reflexivas o exploratorias podrían ayudarte 
        a reconectar con tu esencia.
        """)
    else:
        st.success(f"""
        **✅ Balance Psicológico Funcional**
        
        Tu relación entre Persona ({persona_score:.2f}), Sombra ({sombra_score:.2f}) y Self ({self_score:.2f}) 
        demuestra un buen equilibrio. Estás integrando tus aspectos públicos, internos y tu sentido de propósito 
        de manera relativamente armoniosa.
        """)
    
    # --- GENERAR Y DESCARGAR REPORTES ---
    st.markdown("---")
    st.subheader("📥 Descargar Reportes")
    
    excel_buffer = generar_excel_reporte(respuestas_finales, promedios_roles, promedios_est)
    pdf_buffer = generar_pdf_reporte(respuestas_finales, promedios_roles, promedios_est, fig_radar, fig_bar)
    
    col_down1, col_down2, col_down3 = st.columns(3)
    
    with col_down1:
        st.download_button(
            label="📊 Descargar Excel",
            data=excel_buffer,
            file_name=f"Reporte_Arquetipos_{datetime.now().strftime('%d%m%Y_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    with col_down2:
        st.download_button(
            label="📄 Descargar PDF",
            data=pdf_buffer,
            file_name=f"Reporte_Arquetipos_{datetime.now().strftime('%d%m%Y_%H%M%S')}.pdf",
            mime="application/pdf",
            use_container_width=True
        )
    
    with col_down3:
        if st.button("🔄 Reiniciar Test", use_container_width=True, key="reiniciar_test"):
            st.session_state.clear()
            st.rerun()
    
    st.caption("📌 Nota: Este instrumento sigue principios de análisis de arquetipos Junguianos. Los resultados son tendencias narrativas, no diagnósticos psiquiátricos concluyentes.")